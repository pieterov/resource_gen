# Import module.
import os
import re

import pandas as pd

from datetime             import datetime

from .f_get_root_folder   import f_get_root_folder
from .f_now               import f_now
from .f_var_name          import f_var_name
from .i_variables         import C_PATH_DATA

# Define function.
def f_write_data_to_file(

    l_df,
    c_name = 'Temp',
    c_path = C_PATH_DATA,
    c_type = 'xlsx',
    l_name = None
    ):

    """
    Write object to file.

    Parameters
    ----------
    l_df: 'list' or 'Pandas Series' of values, 'Pandas DataFrame', or 'list' of 'Pandas DataFrame's.
        Data object to write to file.    
    c_name: 'str'
        Name of the file where data object will be saved in.
    c_path: 'str'
        Path where file will be saved.
    c_type: 'str'
        Reference to file type (default: 'xlsx').
    l_name: 'str', or 'list' of 'str'
        Names to be used as sheet names or added to file name (default: 'None').

    Returns
    -------
    -
        Print statement in console to confirm writing of data to file.

    Testing
    -------
    l_df          = [pd.DataFrame({'a': [1,2,3,2,3,3], 'b': [5,6,7,8,9,9]}), pd.DataFrame({'a': [1,2,3], 'b': [5,6,7]})]
    l_df          =  pd.DataFrame({'a': [1,2,3,2,3,3], 'b': [5,6,7,8,9,9]})
    l_df          = pd.Series([1,2,3,4])
    l_df          = [1,2,3,2,3,3]    
    c_name = "Data file"
    c_path        = C_PATH_DELIVERABLES
    c_type        = 'xlsx'
    c_type        = 'csv'
    c_type        = 'parquet'
    l_name        = None
    l_name        = ['DATA1', 'DATA2']

    f_write_data_to_file(l_df, c_name, c_path, c_type, l_name)
    """ 


#----------------------------------------------------------------------------------------------------------------------
# Initialization.
#----------------------------------------------------------------------------------------------------------------------

    # Assign object name, for later use when communicating to user, see end.
    c_l_df = f_var_name(l_df)

    # Valid file types.
    l_type_valid = ['xlsx', 'csv', 'parquet']

    # Current date and time.
    c_now  = f_now() + " - "


    # Check on type of l_df and make corrections as needed.
    if isinstance(l_df, list) and not isinstance(l_df[0], pd.DataFrame):
        l_df = pd.Series(l_df)

    if isinstance(l_df, pd.Series):
        l_df = pd.DataFrame({'l_df': l_df})

    if isinstance(l_df, pd.DataFrame):
        l_df = [l_df]


    # Convert l_name to list if it is a str.
    if isinstance(l_name, str):
        l_name = [l_name]

    # Create l_name if not provided.
    if l_name is None:
        l_name = ['data' + str(i+1) for _,i in enumerate(l_df)]


#----------------------------------------------------------------------------------------------------------------------
# Error check.
#----------------------------------------------------------------------------------------------------------------------

    if len(l_df) != len(l_name):
        raise IndexError("Length of 'l_df' and 'l_name' are not the same.")

    if c_type not in l_type_valid :
        raise ValueError(f"You did not provide a valid file type. Choose 'c_type' to be one of {', '.join(l_type_valid)}.")


#----------------------------------------------------------------------------------------------------------------------
# Main.
#----------------------------------------------------------------------------------------------------------------------

    # Excel - Store dataframe(s) in separate worksheets in same workbook.
    # To check later - https://xlsxwriter.readthedocs.io/example_pandas_table.html
    if c_type == 'xlsx':
  
        from openpyxl import load_workbook
        from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side, Alignment
        
        # with pd.ExcelWriter(os.path.join(c_path, c_now + c_name + "." + c_type)) as writer:

        #     for i in range(len(l_df)):

        #         l_df[i].to_excel(
        #             excel_writer = writer,
        #             sheet_name   = l_name[i],
        #             index        = False
        #         )

        with pd.ExcelWriter(
            
            path   = os.path.join(c_path, c_now + c_name + "." + c_type),
            engine = 'openpyxl',
            mode   = 'w'

        ) as writer:
            
            # Access the workbook
            #workbook = writer.book

            for i in range(len(l_df)):

                l_df[i].to_excel(
                    excel_writer = writer,
                    sheet_name   = l_name[i],
                    index        = False,
                    engine       = 'openpyxl'
                )

                # Access the worksheet.
                worksheet = writer.sheets[l_name[i]]

                # Set the zoom level for the worksheet to 150%.
                worksheet.sheet_view.zoomScale = 150

                # # Set column width for columns C and D.
                # dc_col_width = {'A': 15, 'B': 10, 'C': 50, 'D': 50, 'E': 10, 'F': 10, 'G': 50}

                # for k, v in dc_col_width.items():

                #     # Set the column width for column C and D to width of 85.
                #     worksheet.column_dimensions[f'{k}'].width = v

                #     # # Wrap text in columns.
                #     # for cell in worksheet[f'{col}']:
                #     #     cell.alignment = Alignment(wrapText=True)

                # # Align text in vertical direction.
                # for col in ['A', 'B', 'C', 'D', 'E']:

                #     # Wrap and vertically center text in all columns.
                #     for cell in worksheet[f'{col}']:

                #         cell.alignment = Alignment(
                #             vertical = 'center',
                #             wrapText = True
                #         )

                # # Define the name of the style you want to create or modify
                # style_name = 'TableStyle'

                # # Check if the style already exists in the workbook
                # style_found = False
                # for named_style in workbook.named_styles:
                #     if named_style.name == style_name:

                #         # If the style exists, modify it
                #         table_style = named_style
                #         style_found = True
                #         break

                # if not style_found:
                #     # If the style does not exist, create a new named style
                #     table_style = NamedStyle(name=style_name)

                # # Apply font style
                # table_style.font = Font(
                #     #bold = True,
                #     name = 'Arial',  # Specify the font name (e.g., Arial)
                #     size = 11        # Specify the font size
                # )

                # # Apply fill color.
                # table_style.fill = PatternFill(
                #     start_color = 'FFFF00',
                #     end_color   = 'FFFF00',
                #     fill_type   = 'solid'
                # )

                # # Apply borders
                # border = Border(
                #     left=Side(border_style='thin', color='000000'),
                #     right=Side(border_style='thin', color='000000'),
                #     top=Side(border_style='thin', color='000000'),
                #     bottom=Side(border_style='thin', color='000000')
                # )
                # table_style.border = border

                # # Apply text alignment.
                # table_style.alignment = Alignment(
                #     #horizontal = 'center',
                #     vertical   = 'center',
                #     wrapText   = True
                # )

                # # Apply the style to a range (e.g., the entire table).
                # for row in worksheet.iter_rows(                    
                #     min_row = 2,
                #     max_row = len(l_df[i]) + 1,
                #     min_col = 1,
                #     max_col = len(l_df[i].columns)
                # ):
                #     for cell in row:
                #         cell.style = table_style

                # Save the Excel file
                #writer.save()







                ## I NEED TO MAKE STYLE CHANGES AFTER I SAVED IT
                ## THE SCRIPT BELOW I GOT FROM CHAT GPT

                # pandas.ExcelWriter in the 'openpyxl' engine doesn't provide direct access to named styles to update or modify them.
                # To update styles using pandas.ExcelWriter, you'll typically need to rely on the native Excel styling capabilities 
                # available through openpyxl or by manually modifying the Excel file after it has been generated by pandas.

                # Here's an example of how to update styles after writing an Excel file using pandas.ExcelWriter:

                # import pandas as pd
                # from openpyxl import load_workbook
                # from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

                # # Create a sample DataFrame
                # data = {'Column1': [1, 2, 3], 'Column2': ['A', 'B', 'C']}
                # df = pd.DataFrame(data)

                # # Create an ExcelWriter object using openpyxl engine
                # with pd.ExcelWriter('styled_table.xlsx', engine='openpyxl', mode='w') as writer:
                #     df.to_excel(writer, sheet_name='Sheet1', index=False)

                # # Load the generated Excel file
                # workbook = load_workbook('styled_table.xlsx')

                # # Access a specific sheet within the workbook (replace 'Sheet1' with your sheet name)
                # sheet = workbook['Sheet1']

                # # Access the specific worksheet
                # worksheet = sheet

                # # Define a Font object for font formatting
                # font = Font(
                #     bold=True,
                #     name='Arial',  # Specify the font name (e.g., Arial)
                #     size=12,       # Specify the font size
                # )

                # # Apply the font formatting to a range of cells (e.g., A2:B4)
                # for row in worksheet.iter_rows(
                #     min_row=2,
                #     max_row=4,
                #     min_col=1,
                #     max_col=2
                # ):
                #     for cell in row:
                #         cell.font = font

                # # Define other styles (fill color, borders, alignment, etc.) and apply them as needed

                # # Save the modified workbook
                # workbook.save('styled_table_updated.xlsx')







    # CSV - Store dataframe(s) in separate CSV files.
    if c_type == 'csv':

        for i in range(len(l_df)):

            l_df[i].to_csv(
                path  = os.path.join(c_path, c_now + c_name + " - " + l_name[i] + "." + c_type),
                index = False
            )


    # Parquet - Store dataframe(s) in separate CSV files.
    if c_type == 'parquet':

        for i in range(len(l_df)):

            l_df[i].to_parquet(
                path   = os.path.join(c_path, c_now + c_name + " - " + l_name[i] + "." + c_type),
                index  = False,
                engine = 'pyarrow'
            )


    # l_df[0].iloc[:,:5].to_parquet(c_path + c_now + c_name + " - " + l_name[i] + "." + c_type, index=False)

    # Comms to the user.
    print(f"\nWriting at : {datetime.now()}")

    print(f"Object     : '{c_l_df}'")

    print(f"Name       : '{c_now + c_name + '.' + c_type}'")

    print(f"As         : '{c_type}'")

    print(f"Path       : '.../{re.sub(f_get_root_folder(), '', c_path)}'")

    print(f"==========================")

#----------------------------------------------------------------------------------------------------------------------
# Return results.
#----------------------------------------------------------------------------------------------------------------------
